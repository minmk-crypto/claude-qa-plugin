#!/usr/bin/env python3
"""
QA TC 결과 → 엑셀 + md 리포트 생성

사용법:
  python3 generate-report.py --input results.json --output-dir ./output/reports --name pdp

입력 JSON 형식:
  {
    "meta": {
      "date": "2026-04-09",
      "page": "PDP",
      "urls": ["https://..."],
      "viewports": [{"name":"mobile","w":390,"h":844}, ...]
    },
    "tcs": [
      {
        "id": "TC-A01",
        "priority": "매우높음",
        "category": "구매",
        "scenario": "Add to Cart 버튼 존재",
        "action": "PDP 진입",
        "expected": "버튼 보임, 클릭 가능",
        "auto": true
      }, ...
    ],
    "results": {
      "TC-A01": {
        "peach-glow|390": {"status": "PASS", "note": ""},
        "peach-glow|768": {"status": "PASS", "note": ""},
        "velvet-tint|390": {"status": "PASS", "note": ""},
        ...
      }, ...
    },
    "screenshots": {
      "TC-A03|lucky-shine|390": "screenshots/tc-a03-soldout.png"
    }
  }
"""

import json
import sys
import argparse
from datetime import date
from pathlib import Path

def main():
    parser = argparse.ArgumentParser(description='QA 리포트 생성')
    parser.add_argument('--input', required=True, help='결과 JSON 파일')
    parser.add_argument('--output-dir', default='./output/reports', help='출력 디렉토리')
    parser.add_argument('--name', default='qa', help='페이지명 (파일명에 사용)')
    args = parser.parse_args()

    with open(args.input) as f:
        data = json.load(f)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    today = data.get('meta', {}).get('date', date.today().isoformat())
    page_name = args.name
    tcs = data.get('tcs', [])
    results = data.get('results', {})
    screenshots = data.get('screenshots', {})
    urls = data.get('meta', {}).get('urls', [])
    viewports = data.get('meta', {}).get('viewports', [])

    # --- 엑셀 생성 ---
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'QA Test Cases'

        # 헤더
        base_headers = ['TC ID', '우선순위', '카테고리', '시나리오', '동작', '기대결과', '자동화']

        # URL × 해상도 컬럼
        result_headers = []
        for url in urls:
            short = url.split('/')[-1][:15] if '/' in url else url[:15]
            for vp in viewports:
                result_headers.append(f'{short}\n({vp["w"]})')

        headers = base_headers + result_headers + ['비고']

        # 스타일
        hfill = PatternFill(start_color='333F4F', end_color='333F4F', fill_type='solid')
        hfont = Font(bold=True, color='FFFFFF', size=9)
        fills = {
            'PASS': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'FAIL': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'SKIP': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
            'N/A': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
            '수동': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
            '확인필요': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
        }
        prio_fills = {
            '매우높음': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            '높음': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
            '보통': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
            '낮음': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        }
        tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

        # 헤더 행
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill
            c.font = hfont
            c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
            c.border = tb

        # TC 행
        for i, tc in enumerate(tcs, 2):
            tc_id = tc['id']
            vals = [tc_id, tc['priority'], tc['category'], tc['scenario'],
                    tc['action'], tc['expected'], 'Y' if tc.get('auto') else '수동']

            for j, v in enumerate(vals, 1):
                c = ws.cell(row=i, column=j, value=v)
                c.border = tb
                c.alignment = Alignment(wrap_text=True, vertical='center')

            # 우선순위 색상
            ws.cell(row=i, column=2).fill = prio_fills.get(tc['priority'], PatternFill())

            # 결과 셀
            col_idx = len(base_headers) + 1
            tc_results = results.get(tc_id, {})
            notes = []

            for url in urls:
                short = url.split('/')[-1][:15] if '/' in url else url[:15]
                for vp in viewports:
                    key = f'{short}|{vp["w"]}'
                    r = tc_results.get(key, {})
                    status = r.get('status', '')
                    note = r.get('note', '')

                    c = ws.cell(row=i, column=col_idx, value=status)
                    c.border = tb
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    if status in fills:
                        c.fill = fills[status]

                    if note:
                        notes.append(f'[{key}] {note}')
                    col_idx += 1

            # 비고
            ws.cell(row=i, column=col_idx, value='\n'.join(notes)).border = tb

        # 열 너비
        widths = [9, 8, 8, 28, 22, 28, 5] + [10] * len(result_headers) + [35]
        for i, w in enumerate(widths, 1):
            if i <= len(headers):
                ws.column_dimensions[get_column_letter(i)].width = w

        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(tcs)+1}'
        ws.freeze_panes = f'{get_column_letter(len(base_headers)+1)}2'
        ws.row_dimensions[1].height = 40

        xlsx_path = output_dir / f'{today}-qa-{page_name}.xlsx'
        wb.save(xlsx_path)
        print(f'Excel: {xlsx_path}')

    except ImportError:
        print('openpyxl 미설치 — 엑셀 생성 건너뜀', file=sys.stderr)
        xlsx_path = None

    # --- md 리포트 ---
    total = 0
    counts = {'PASS': 0, 'FAIL': 0, 'SKIP': 0, 'N/A': 0, '수동': 0, '확인필요': 0}
    fail_list = []

    for tc in tcs:
        tc_id = tc['id']
        tc_results = results.get(tc_id, {})
        for key, r in tc_results.items():
            status = r.get('status', '')
            total += 1
            counts[status] = counts.get(status, 0) + 1
            if status == 'FAIL':
                fail_list.append({
                    'tc': tc_id,
                    'priority': tc['priority'],
                    'scenario': tc['scenario'],
                    'key': key,
                    'note': r.get('note', ''),
                    'screenshot': screenshots.get(f'{tc_id}|{key}', '')
                })

    md_lines = [
        f'---',
        f'type: output',
        f'output_type: report',
        f'date_created: {today}',
        f'tags: [qa, {page_name}]',
        f'---',
        f'',
        f'# QA 검증 리포트 — {today} ({page_name.upper()})',
        f'',
        f'## 요약',
        f'',
        f'| 항목 | 값 |',
        f'|------|-----|',
        f'| 총 검증 건수 | {total} |',
        f'| PASS | {counts.get("PASS", 0)} |',
        f'| **FAIL** | **{counts.get("FAIL", 0)}** |',
        f'| 확인필요 | {counts.get("확인필요", 0)} |',
        f'| SKIP | {counts.get("SKIP", 0)} |',
        f'| N/A | {counts.get("N/A", 0)} |',
        f'| 수동 | {counts.get("수동", 0)} |',
        f'',
    ]

    if fail_list:
        md_lines.extend([
            f'## FAIL 항목',
            f'',
            f'| TC | 우선순위 | 시나리오 | 상품/해상도 | 설명 |',
            f'|---|---|---|---|---|',
        ])
        for f in fail_list:
            screenshot_link = f' ![screenshot]({f["screenshot"]})' if f['screenshot'] else ''
            md_lines.append(f'| {f["tc"]} | {f["priority"]} | {f["scenario"]} | {f["key"]} | {f["note"]}{screenshot_link} |')
        md_lines.append('')

    md_path = output_dir / f'{today}-qa-{page_name}.md'
    md_path.write_text('\n'.join(md_lines), encoding='utf-8')
    print(f'Report: {md_path}')

    # --- 요약 출력 ---
    print(f'\n=== QA 결과 요약 ===')
    print(f'PASS: {counts.get("PASS",0)} | FAIL: {counts.get("FAIL",0)} | SKIP: {counts.get("SKIP",0)}')
    if fail_list:
        print(f'\nFAIL {len(fail_list)}건:')
        for f in fail_list:
            print(f'  [{f["priority"]}] {f["tc"]}: {f["scenario"]} — {f["key"]} — {f["note"]}')

if __name__ == '__main__':
    main()
